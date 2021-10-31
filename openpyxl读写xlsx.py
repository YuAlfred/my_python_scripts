# -*- encoding: utf-8 -*-
"""
@Author      : Alfred T
@Time        : 2021/10/31 12:56
@description : 
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

import warnings

warnings.filterwarnings('ignore')

# openpyxl只能处理 .xlsx 合适的表格
# 加载工作簿
wb = load_workbook('./成都市郫都区绵实外国语学校(十二年一贯制学校).xlsx')
ws = wb.get_sheet_by_name("教基1001_十二年一贯制学校")


ans = load_workbook('./ans.xlsx')
ansSheet = ans.get_sheet_by_name("Sheet1")

print(ws.cell(3, 3).value)

ansSheet.cell(1, 1).value = ws.cell(3, 3).value
print(ansSheet.cell(1, 1).value)
ans.save("./ans.xlsx")
