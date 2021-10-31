# -*- encoding: utf-8 -*-
"""
@Author      : Alfred T
@Time        : 2021/10/31 15:28
@description :
"""
import os
from openpyxl import Workbook, load_workbook, worksheet
import warnings

warnings.filterwarnings('ignore')


# 统计小学
# path：excel路径
# suffix：子表后缀
# ws: 结果表worksheet 子表引用
# index：当前写入行数
def count_primary_school(path, suffix, ws: worksheet, index):
    # 先打开此表
    cur_work_book = load_workbook(path)

    # 教基1001信息写入
    cur_work_sheet = cur_work_book.get_sheet_by_name("教基1001_" + suffix)
    ws.cell(index, 1).value = cur_work_sheet.cell(3, 3).value
    ws.cell(index, 2).value = cur_work_sheet.cell(4, 3).value
    ws.cell(index, 3).value = cur_work_sheet.cell(14, 3).value
    ws.cell(index, 4).value = cur_work_sheet.cell(27, 3).value
    ws.cell(index, 5).value = cur_work_sheet.cell(28, 3).value

    # 教基1102信息写入
    cur_work_sheet = cur_work_book.get_sheet_by_name("教基1102_" + suffix)
    ws.cell(index, 6).value = cur_work_sheet.cell(8, 3).value
    ws.cell(index, 7).value = cur_work_sheet.cell(9, 3).value
    ws.cell(index, 8).value = cur_work_sheet.cell(10, 3).value
    ws.cell(index, 9).value = cur_work_sheet.cell(11, 3).value
    ws.cell(index, 10).value = cur_work_sheet.cell(12, 3).value

    # 教基1102信息写入
    cur_work_sheet = cur_work_book.get_sheet_by_name("教基1102_" + suffix)
    ws.cell(index, 11).value = cur_work_sheet.cell(8, 3).value
    ws.cell(index, 12).value = cur_work_sheet.cell(9, 3).value
    ws.cell(index, 13).value = cur_work_sheet.cell(10, 3).value
    ws.cell(index, 14).value = cur_work_sheet.cell(11, 3).value
    ws.cell(index, 15).value = cur_work_sheet.cell(12, 3).value


#  主程序
if __name__ == "__main__":
    ans_work_book = load_workbook("./v2.xlsx")

    # 先统计小学
    ws = ans_work_book.get_sheet_by_name("基教小学")
    index = 7
    root_dir = "./中小学/小学/"
    for f in os.listdir(root_dir):
        if os.path.isfile(root_dir + f) and f[0] != "~":
            print("正在处理：" + f)
            count_primary_school(root_dir + f, f[f.index("(") + 1:f.index(")")], ws, index)
            index += 1

    ans_work_book.save("./v2.xlsx")
